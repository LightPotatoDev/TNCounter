import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill

ABSPATH = os.path.abspath(os.path.dirname(__file__))
ITEMS_RES_PATH = ''.join((os.path.join(ABSPATH, '..'), '\\resources\\items'))

def make_tactical_sheet(item_df, extra_df):
    excel_file = 'TNCount.xlsx'
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        item_df.to_excel(writer, sheet_name = 'Item Count', index=True)
        extra_df.to_excel(writer, sheet_name = 'Extra Info', index = True)
    
    workbook = load_workbook(excel_file)
    sheet = workbook.active
    
    #insert images at the top row
    for i in range(2,sheet.max_column+1):
        cell = sheet.cell(row=1, column=i)
        item_name = cell.value
        item_img_path = ''.join((ITEMS_RES_PATH, '\\', item_name, '.png'))
        if os.path.isfile(item_img_path):
            sheet.add_image(Image(item_img_path), cell.coordinate)
        cell.value = ''
        
    #change the color of 'Total / Medal room' cells
    total_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    medal_fill = PatternFill(start_color='EFEFEF', end_color='EFEFEF', fill_type='solid')
    for i in range(2,sheet.max_row+1):
        cell = sheet.cell(row=i, column=1)
        if cell.value == 'Total':
            for j in range(1,sheet.max_column+1):
                color_cell = sheet.cell(row=i, column=j)
                color_cell.fill = total_fill
        if cell.value[0] == 'M':
            for j in range(1,sheet.max_column+1):
                color_cell = sheet.cell(row=i, column=j)
                color_cell.fill = medal_fill
                
    workbook.save(excel_file)