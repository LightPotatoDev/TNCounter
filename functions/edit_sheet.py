import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from . import img_process

ABSPATH = os.path.abspath(os.path.dirname(__file__))
ITEMS_RES_PATH = ''.join((os.path.join(ABSPATH, '..'), '\\resources\\items'))

def edit_tactical_sheet(excel_file):
    workbook = load_workbook(excel_file)
    sheet = workbook.active
    
    for i in range(2,sheet.max_column+1):
        cell = sheet.cell(row=1, column=i)
        item_name = cell.value
        item_img_path = ''.join((ITEMS_RES_PATH, '\\', item_name, '.png'))
        if os.path.isfile(item_img_path):
            sheet.add_image(Image(item_img_path), cell.coordinate)
    
    workbook.save(excel_file)